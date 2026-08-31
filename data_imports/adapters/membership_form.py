from io import BytesIO

from openpyxl import load_workbook

from data_imports.services.normalization import json_safe


SHEET_NAME = "Form Responses 1"
REQUIRED_HEADERS = (
    "Timestamp", "First Name ", "Last Name ", "Gender", "Age ",
    "Email (preferably your personal email)", "Mobile Number", "Location",
    "What industry do you work in?", "What is your current role or job title?", "Linkedin URL",
)


class MembershipFormStructureError(Exception):
    pass


def iter_membership_form_rows(workbook_bytes):
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise MembershipFormStructureError("Required Membership Form worksheet is missing.")
    worksheet = workbook[SHEET_NAME]
    headers = tuple(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    missing = [header for header in REQUIRED_HEADERS if header not in headers]
    if missing:
        raise MembershipFormStructureError("Membership Form workbook is missing required headers: " + ", ".join(missing))
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or (isinstance(value, str) and not value.strip()) for value in values):
            continue
        yield row_number, {header: json_safe(value) for header, value in zip(headers, values)}
