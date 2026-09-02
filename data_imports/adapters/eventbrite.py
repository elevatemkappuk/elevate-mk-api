from io import BytesIO

from openpyxl import load_workbook

from data_imports.services.normalization import json_safe


REQUIRED_COLUMN_ALIASES = {
    "buyer_first_name": ("Buyer First Name",),
    "buyer_last_name": ("Buyer Surname", "Buyer Last Name"),
    "buyer_email": ("Buyer Email",),
    "mobile": ("Phone Number",),
    "city": ("Purchaser Town/City", "Purchaser Town", "Purchaser City"),
    "county": ("Purchaser County",),
    "country": ("Purchaser Country",),
    "external_event_id": ("Event ID",),
    "event_name": ("Event Name",),
    "event_start_date": ("Event Start Date",),
    "event_start_time": ("Event Start Time",),
    "event_timezone": ("Event Timezone", "Event Time Zone"),
    "event_location": ("Event Location",),
    "external_order_id": ("Order ID",),
    "order_date": ("Order Date",),
    "ticket_quantity": ("Ticket Quantity",),
    "guest": ("Guest",),
}


class EventbriteStructureError(Exception):
    pass


def iter_eventbrite_rows(workbook_bytes):
    workbook = load_workbook(BytesIO(workbook_bytes), read_only=True, data_only=True)
    worksheet = workbook.active
    headers = tuple(next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True)))
    header_map = resolve_header_map(headers)
    for row_number, values in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        if _is_empty(values) or _is_totals_row(values):
            continue
        yield row_number, {header: json_safe(value) for header, value in zip(headers, values)}, header_map


def resolve_header_map(headers):
    normalized_headers = {_canonical_header(header): header for header in headers if header is not None}
    header_map = {}
    missing = []
    for field, aliases in REQUIRED_COLUMN_ALIASES.items():
        matched_header = next((normalized_headers.get(_canonical_header(alias)) for alias in aliases), None)
        if matched_header is None:
            missing.append(field)
        else:
            header_map[field] = matched_header
    if missing:
        raise EventbriteStructureError("Eventbrite workbook is missing required columns.")
    return header_map


def _canonical_header(value):
    return " ".join(str(value or "").strip().casefold().split())


def _is_empty(values):
    return all(value is None or (isinstance(value, str) and not value.strip()) for value in values)


def _is_totals_row(values):
    return any(isinstance(value, str) and value.strip().casefold() == "totals" for value in values)
